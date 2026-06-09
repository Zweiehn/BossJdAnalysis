"""飞书多维表格同步模块 - 将 Excel 数据推送到飞书 Bitable"""

import json
import time
import requests
from datetime import datetime
from models import FIELD_ORDER
from config import Config


# 飞书 API 端点
FEISHU_BASE = "https://open.feishu.cn"
AUTH_URL = f"{FEISHU_BASE}/open-apis/auth/v3/tenant_access_token/internal"
BITABLE_BATCH_CREATE = (
    f"{FEISHU_BASE}/open-apis/bitable/v1/apps/{{app_token}}/tables/{{table_id}}/records/batch_create"
)
BITABLE_FIELDS = (
    f"{FEISHU_BASE}/open-apis/bitable/v1/apps/{{app_token}}/tables/{{table_id}}/fields"
)
BITABLE_LIST = (
    f"{FEISHU_BASE}/open-apis/bitable/v1/apps/{{app_token}}/tables/{{table_id}}/records"
)

# 飞书字段类型对照
FIELD_TYPE_MAP = {
    1: "文本", 2: "数字", 3: "单选", 4: "多选", 5: "日期",
    7: "复选框", 11: "电话", 13: "超链接", 15: "附件",
    17: "关联", 1001: "多行文本", 1003: "条码",
}


class FeishuSyncer:
    """飞书多维表格同步器"""

    def __init__(self, config: Config = None):
        self.config = config or Config()
        self._token = None
        self._token_expires_at = 0

    # ──────────── 配置检查 ────────────

    def is_configured(self) -> bool:
        """检查飞书配置是否完整"""
        return bool(
            self.config.feishu_app_id
            and self.config.feishu_app_secret
            and self.config.feishu_base_token
            and self.config.feishu_table_id
        )

    def check_config(self) -> str:
        """返回配置检查结果，空字符串表示就绪"""
        if not self.config.feishu_app_id:
            return "缺少 App ID"
        if not self.config.feishu_app_secret:
            return "缺少 App Secret"
        if not self.config.feishu_base_token:
            return "缺少 Base Token（多维表格URL中的ID）"
        if not self.config.feishu_table_id:
            return "缺少 Table ID"
        return ""

    # ──────────── Token 管理 ────────────

    def _ensure_token(self):
        """获取或刷新 tenant_access_token"""
        if time.time() < self._token_expires_at - 60:
            return

        resp = requests.post(
            AUTH_URL,
            json={
                "app_id": self.config.feishu_app_id,
                "app_secret": self.config.feishu_app_secret,
            },
            timeout=10,
        )
        if resp.status_code != 200:
            raise RuntimeError(f"飞书认证失败 (HTTP {resp.status_code})")

        data = resp.json()
        if data.get("code") != 0:
            msg = data.get("msg", "未知错误")
            raise RuntimeError(f"飞书认证失败: {msg}")

        self._token = data["tenant_access_token"]
        self._token_expires_at = time.time() + data.get("expire", 7200)

    def _get_headers(self) -> dict:
        self._ensure_token()
        return {
            "Authorization": f"Bearer {self._token}",
            "Content-Type": "application/json",
        }

    # ──────────── 字段类型查询 ────────────

    def get_field_schema(self) -> dict:
        """
        获取飞书多维表格的字段元数据

        Returns:
            { "字段名": {"type": 4, "type_name": "多选"}, ... }
        """
        url = BITABLE_FIELDS.format(
            app_token=self.config.feishu_base_token,
            table_id=self.config.feishu_table_id,
        )
        resp = requests.get(url, headers=self._get_headers(), timeout=15)
        if resp.status_code != 200:
            raise RuntimeError(f"获取字段信息失败 (HTTP {resp.status_code})")

        data = resp.json()
        if data.get("code") != 0:
            raise RuntimeError(f"获取字段信息失败: {data.get('msg', '未知错误')}")

        schema = {}
        for item in data.get("data", {}).get("items", []):
            field_name = item.get("field_name", "")
            field_type = item.get("type", 0)
            schema[field_name] = {
                "type": field_type,
                "type_name": FIELD_TYPE_MAP.get(field_type, f"未知({field_type})"),
            }
        return schema

    def _format_value(self, value, field_type: int):
        """
        根据飞书字段类型格式化值

        多选(4) → ["是"]   单选(3) → "是"
        复选框(7) → True/False
        数字(2) → 数值
        """
        if value is None or value == "":
            return None

        if field_type == 4:   # 多选 → 数组
            if isinstance(value, list):
                return value
            return [str(value)]

        if field_type == 3:   # 单选 → 字符串
            return str(value)

        if field_type == 7:   # 复选框
            if isinstance(value, bool):
                return value
            return str(value) in ("是", "true", "True", "1", "yes")

        if field_type == 2:   # 数字
            try:
                return float(value)
            except (ValueError, TypeError):
                return str(value)

        if field_type == 5:   # 日期
            return str(value)

        return str(value)      # 文本等

    # ──────────── 核心同步 ────────────

    def sync_excel_to_bitable(self, excel_path: str = None) -> dict:
        """
        将 Excel 中的数据同步到飞书多维表格

        Args:
            excel_path: Excel 文件路径，默认从配置读取

        Returns:
            {"success": N, "failed": N, "errors": [...]}
        """
        from openpyxl import load_workbook

        if not self.is_configured():
            missing = self.check_config()
            raise RuntimeError(f"飞书配置不完整: {missing}")

        if not excel_path:
            excel_path = self.config.resolve_excel_path()
        if not excel_path or not __import__("os").path.exists(excel_path):
            raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")

        # 读取 Excel 数据
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            wb.close()
            return {"success": 0, "failed": 0, "errors": [], "info": "Excel 中没有数据行"}

        headers = list(rows[0])  # 表头
        data_rows = rows[1:]      # 数据行

        # 查询飞书表格字段类型，用于格式化数据
        try:
            field_schema = self.get_field_schema()
            matched = sum(1 for h in headers if h and h.strip() in field_schema)
            print(f"飞书字段匹配: {matched}/{len([h for h in headers if h])}")
        except Exception as e:
            print(f"获取字段类型失败，按默认格式发送: {e}")
            field_schema = {}

        # 构建飞书记录（按字段类型格式化值）
        records = []
        for row in data_rows:
            fields = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i] and val is not None:
                    field_name = str(headers[i]).strip()
                    ft = field_schema.get(field_name, {}).get("type", 0)
                    formatted = self._format_value(val, ft)
                    if formatted is not None:
                        fields[field_name] = formatted
            if fields:
                records.append({"fields": fields})

        wb.close()

        if not records:
            return {"success": 0, "failed": 0, "errors": [], "info": "没有有效数据可同步"}

        # 去重：查询飞书已有记录，只同步新增的
        try:
            existing_fps = self._get_existing_fingerprints()
            if existing_fps:
                before = len(records)
                records = self._filter_new_records(records, existing_fps)
                skipped = before - len(records)
                print(f"去重: 跳过 {skipped} 条，待同步 {len(records)} 条")
            else:
                print(f"飞书表格为空，将同步全部 {len(records)} 条")
        except Exception as e:
            print(f"去重查询失败（将同步全部）: {e}")

        if not records:
            return {"success": 0, "failed": 0, "errors": [], "info": "所有数据已在飞书中，无需同步"}

        # 分批写入（每批最多 1000 条）
        BATCH_SIZE = 500
        total_success = 0
        total_failed = 0
        all_errors = []

        for start in range(0, len(records), BATCH_SIZE):
            batch = records[start:start + BATCH_SIZE]
            result = self._batch_create(batch, start)
            total_success += result["success"]
            total_failed += result["failed"]
            all_errors.extend(result["errors"])

        return {
            "success": total_success,
            "failed": total_failed,
            "errors": all_errors,
            "info": f"同步完成: 成功 {total_success} 条"
            + (f", 失败 {total_failed} 条" if total_failed else ""),
        }

    def _batch_create(self, records: list, offset: int = 0) -> dict:
        """批量创建飞书记录"""
        url = BITABLE_BATCH_CREATE.format(
            app_token=self.config.feishu_base_token,
            table_id=self.config.feishu_table_id,
        )

        payload = {"records": records}

        resp = requests.post(
            url,
            headers=self._get_headers(),
            json=payload,
            timeout=30,
        )

        if resp.status_code != 200:
            detail = resp.text[:300]
            return {"success": 0, "failed": len(records),
                    "errors": [(offset, f"HTTP {resp.status_code}: {detail}")]}

        data = resp.json()
        code = data.get("code", -1)

        if code == 0:
            created = len(data.get("data", {}).get("records", []))
            return {"success": created, "failed": len(records) - created, "errors": []}
        else:
            msg = data.get("msg", "未知错误")
            err_msg = self._parse_error(code, msg)
            return {"success": 0, "failed": len(records),
                    "errors": [(offset, err_msg)]}

    @staticmethod
    def _parse_error(code: int, msg: str) -> str:
        """将 API 错误码转为可读信息"""
        guides = {
            91403: "权限不足，请检查：\n"
                   "1. 应用是否已开通 bitable:app 权限\n"
                   "2. 应用是否已发布（创建版本 → 审核通过）\n"
                   "3. Base Token / Table ID 是否正确\n"
                   "4. 多维表格是否已授权给此应用",
            91003: "Base Token 或 Table ID 无效，请检查配置",
            1254104: "单次写入超过1000条限制",
            99999: "飞书服务内部错误，稍后重试",
        }
        guide = guides.get(code)
        if guide:
            return f"错误({code}): {msg}\n\n{guide}"
        return f"错误({code}): {msg}"

    # ──────────── 验证连接 ────────────

    def test_connection(self) -> str:
        """测试飞书连接是否正常，返回空字符串表示成功"""
        try:
            self._ensure_token()
            # 用 list 接口测试权限（GET /records?page_size=1）
            test_url = BITABLE_BATCH_CREATE.format(
                app_token=self.config.feishu_base_token,
                table_id=self.config.feishu_table_id,
            ).replace("/records/batch_create", "/records?page_size=1")

            resp = requests.get(test_url, headers=self._get_headers(), timeout=10)

            if resp.status_code == 200:
                return ""

            data = resp.json()
            code = data.get("code", -1)
            msg = data.get("msg", "未知错误")
            return self._parse_error(code, msg)

        except requests.exceptions.ConnectionError:
            return "无法连接到飞书服务器，请检查网络"
        except requests.exceptions.Timeout:
            return "连接飞书超时"
        except Exception as e:
            return str(e)

    # ──────────── 去重逻辑 ────────────

    # 用作去重指纹的关键字段
    DEDUP_KEYS = ["公司名称", "岗位名称", "录入时间"]

    def _make_fingerprint(self, fields: dict) -> str:
        """从记录字段生成唯一指纹"""
        parts = []
        for key in self.DEDUP_KEYS:
            val = fields.get(key)
            if val is None:
                val = ""
            if isinstance(val, list):
                val = "|".join(str(v) for v in val)
            parts.append(str(val).strip())
        return "||".join(parts)

    def _get_existing_fingerprints(self) -> set:
        """
        查询飞书多维表格中已有的记录，返回指纹集合

        Returns:
            {"公司A||NLP工程师||2026-06-08", ...}
        """
        fingerprints = set()
        page_token = None
        url_template = BITABLE_LIST.format(
            app_token=self.config.feishu_base_token,
            table_id=self.config.feishu_table_id,
        )

        while True:
            url = url_template + "?page_size=500"
            if page_token:
                url += f"&page_token={page_token}"

            resp = requests.get(url, headers=self._get_headers(), timeout=15)
            if resp.status_code != 200:
                break

            data = resp.json()
            if data.get("code") != 0:
                break

            items = data.get("data", {}).get("items", [])
            for item in items:
                fields = item.get("fields", {})
                fp = self._make_fingerprint(fields)
                fingerprints.add(fp)

            # 检查是否有下一页
            page_token = data.get("data", {}).get("page_token")
            if not page_token:
                break

        return fingerprints

    def _filter_new_records(self, records: list, existing_fps: set) -> list:
        """只保留飞书中尚不存在的记录"""
        new_records = []
        skipped = 0
        for rec in records:
            fp = self._make_fingerprint(rec["fields"])
            if fp in existing_fps:
                skipped += 1
            else:
                new_records.append(rec)
        if skipped:
            print(f"跳过 {skipped} 条已存在的记录")
        return new_records
