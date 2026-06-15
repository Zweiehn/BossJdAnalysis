"""Excel 写入模块 - 自动创建/追加 Excel 数据库"""

import os
import threading
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from models import JDRecord, FIELD_ORDER
from config import Config, get_app_dir


class ExcelWriter:
    """Excel 文件写入器，支持自动创建和追加"""

    def __init__(self, config: Config = None):
        self.config = config or Config()
        self._lock = threading.Lock()  # 线程安全

    def _get_excel_path(self) -> str:
        """获取 Excel 文件路径"""
        path = self.config.excel_path
        if os.path.isabs(path):
            return path
        return os.path.join(get_app_dir(), path)

    _NARROW_FIELDS = {"原始JD文本": 35}

    def _get_headers(self, ws) -> list:
        """读取 Excel 当前的表头行"""
        headers = []
        for col in range(1, ws.max_column + 1):
            h = ws.cell(1, col).value
            if h is not None:
                headers.append(str(h).strip())
        return headers

    def _ensure_headers(self, ws):
        """首次创建：写入完整表头"""
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, field_name in enumerate(FIELD_ORDER, 1):
            cell = ws.cell(1, col_idx, field_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    def _auto_column_width(self, ws, headers: list):
        """自动调整列宽"""
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)

            # 长文本列固定窄宽
            if header in self._NARROW_FIELDS:
                ws.column_dimensions[col_letter].width = self._NARROW_FIELDS[header]
                continue

            max_length = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
                for cell in row:
                    if cell.value:
                        cell_len = sum(2 if '一' <= c <= '鿿' else 1 for c in str(cell.value))
                        max_length = max(max_length, cell_len)

            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 8), 45)

    def _record_to_row(self, record: JDRecord, headers: list) -> list:
        """将 JDRecord 的值按 Excel 表头顺序映射，自定义字段留空"""
        record_dict = {}
        for f in FIELD_ORDER:
            record_dict[f] = getattr(record, f, "")
        return [record_dict.get(h, "") for h in headers]

    def append_record(self, record: JDRecord):
        """
        追加一条记录到 Excel（动态映射表头，支持用户自定义字段）

        Args:
            record: JDRecord 对象
        """
        with self._lock:
            excel_path = self._get_excel_path()
            file_exists = os.path.exists(excel_path)

            if file_exists:
                try:
                    wb = load_workbook(excel_path)
                    ws = wb.active
                except PermissionError:
                    raise PermissionError(f"无法写入 Excel，请先关闭文件: {excel_path}")
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "JD记录"
                self._ensure_headers(ws)

            headers = self._get_headers(ws)
            row_data = self._record_to_row(record, headers)
            ws.append(row_data)

            # 设置新行对齐 + 超链接
            new_row = ws.max_row
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(new_row, col_idx)
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=(header not in self._NARROW_FIELDS),
                )
                # 截图文件列设超链接
                if header == "截图文件" and cell.value:
                    cell.hyperlink = str(cell.value)

            self._auto_column_width(ws, headers)
            ws.freeze_panes = "A2"

            try:
                wb.save(excel_path)
                print(f"数据已写入: {excel_path} (第{new_row - 1}条)")
            except PermissionError:
                raise PermissionError(f"保存 Excel 失败，请先关闭文件: {excel_path}")

    def append_records(self, records: list):
        """批量追加记录"""
        for record in records:
            self.append_record(record)

    def get_record_count(self) -> int:
        """获取当前记录数"""
        excel_path = self._get_excel_path()
        if not os.path.exists(excel_path):
            return 0
        try:
            wb = load_workbook(excel_path, read_only=True)
            ws = wb.active
            count = max(0, ws.max_row - 1)  # 减去表头
            wb.close()
            return count
        except Exception:
            return 0

    def open_excel(self):
        """打开 Excel 文件（系统默认程序）"""
        import subprocess
        excel_path = self._get_excel_path()
        if os.path.exists(excel_path):
            os.startfile(excel_path)
        else:
            raise FileNotFoundError(f"Excel 文件尚未创建: {excel_path}")
